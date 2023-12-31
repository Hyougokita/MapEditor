import tkinter
import tkinter as tk
import tkinter.ttk as ttk
import  csv
import  openpyxl
from openpyxl import  Workbook
from tkinter import  filedialog
from tkinter import simpledialog



class Mapedit(tk.Frame):
    def __init__(self, master = None):
        super().__init__(master)

        #ウィンドウズのサイズ
        self.editor_size_x = 960
        self.editor_size_y = 540


        self.master.title("縁")    # ウィンドウタイトル
        self.master.geometry("960x540") # ウィンドウサイズ(幅x高さ)
        # Canvasの作成
        self.canvas = tk.Canvas(self.master,width = self.editor_size_x,height = self.editor_size_y,bg = "white")
        # Canvasを配置

        # 初めてのマス位置
        self.startX = 10
        self.startY = 20

        # マス
        self.scale = 1.0
        self.masuWidth = 20 * self.scale
        self.masuHeight = 20 * self.scale
        self.padding = 10 * self.scale

        self.masu_x = -1
        self.masu_y = -1

        #建物のフラグを保存するリスト
        self.map_chip = []
        #マス描画用四角形保存リスト
        self.masu_rect_list = []

        #Mouse
        self.mousePosX = 0
        self.mousePosY = 0

        #現在の建築フラグ
        self.constriction_flag = -1

        #サブメニュー関連
        self.sub_menu_active = True
        self.sub_menu_item_number = 0
        self.sub_menu_item_list = []
        self.sub_menu_item_color = ["LightPink","Purple","Gold","Aqua","Azure",
                                    "Brown","Cyan","Green","IndianRed","LightBlue",
                                    "Maroon","Navy","Orange","Salmon","SkyBlue",
                                    "Tan","Tomato","Violet","Wheat","Yellow"]


        self.sub_menu_item_color_can_be_choice = []
        for i in self.sub_menu_item_color:
            self.sub_menu_item_color_can_be_choice.append(i)
        self.sub_menu_button_list = []
        self.sub_menu_button_start_position_x = 820
        self.sub_menu_button_start_position_y = 25
        self.sub_menu_button_padding_y = 45
        self.sub_menu_item_dict = []


        #ファイル名前関連
        self.map_excel_file_full_name = ""
        self.map_excel_file_name = ""

        #map chip移動
        self.RightPressPos_x = 0
        self.RightPressPos_y = 0


    def DrawMasu(self,x,y,color,list):
        self.masu = self.canvas.create_rectangle(x,y,x + self.masuWidth,y + self.masuHeight,fill=color)
        print("type",type(self.masu))
        list.append(self.masu)

    def SetMasu(self,yoko,tate):
        self.masu_rect_list = []
        for i in range(0,yoko):
            temp_list = []
            for j in range(0,tate):
                if __debug__:
                    print("map_chip:",int(self.map_chip[i][j]))
                self.DrawMasu(self.startX + (self.masuWidth + self.padding) * i,
                              self.startY + (self.masuHeight + self.padding) * j,
                              self.sub_menu_item_color[int(self.map_chip[i][j])],temp_list)
            self.masu_rect_list.append(temp_list)
        if __debug__:
            print(self.masu_rect_list)
        self.canvas.pack()

    def CheckMasu(self):
        self.tempMasuX = (int)((self.mousePosX - self.startX) / (self.masuWidth + self.padding))
        self.tempMasuY = (int)((self.mousePosY - self.startY) / (self.masuHeight + self.padding))
        print("(",self.tempMasuX,",",self.tempMasuY,")")

        return  (self.tempMasuX,self.tempMasuY)

    def CheckIn(self,x,y):
        self.tempMasuX,self.tempMasuY = self.CheckMasu()
        if( x > self.startX + (self.masuWidth + self.padding) * self.tempMasuX and
            x < self.startX + (self.masuWidth + self.padding) * self.tempMasuX + self.masuWidth and
            y > self.startY + (self.masuHeight + self.padding) * self.tempMasuY and
            y < self.startY + (self.masuHeight + self.padding) * self.tempMasuY + self.masuHeight):
            print("in")
            return True
        print("no")
        return False

    def ChangeColor(self,x,y):
        if __debug__:
            print("sub_menu_item_color",self.sub_menu_item_color)
            print("sub_menu_item_dict",self.sub_menu_item_dict)
        self.canvas.itemconfig(self.masu_rect_list[x][y],fill = self.sub_menu_item_color[self.sub_menu_item_dict[self.constriction_flag]['obj_flag'] - 1])


    def WriteCsv(self,col,cow,constriction_flag):
        wb = Workbook()
        ws = wb.active
        ws.cell(col + 1,cow + 1,constriction_flag)
        wb.save(r'map1.xlsx')
        pass

    def SaveMapChipToExcel(self):
        if (self.map_excel_file_full_name != ""):
            book = openpyxl.load_workbook(self.map_excel_file_full_name)

            # マップチップの保存
            sheet = book.get_sheet_by_name("Map")
            for i in range(self.masu_x):
                for j in range(self.masu_y):
                    sheet.cell(row = j + 1, column = i + 1).value = self.map_chip[i][j]

            # マップチップのサイズの保存
            sheet = book.get_sheet_by_name("MapInfo")
            sheet.cell(row=1, column=2).value = self.masu_x
            sheet.cell(row=2, column=2).value = self.masu_y

            # Excelファイルを保存して閉じる
            book.save(self.map_excel_file_full_name)

            if __debug__:
                print("保存しました！")

    def LeftClickEvent(self,event):
        print("X:",event.x," Y:",event.y)
        self.mousePosX = event.x
        self.mousePosY = event.y
        if(self.CheckIn(event.x,event.y)):
            self.ChangeColor(self.tempMasuX,self.tempMasuY)
            self.map_chip[self.tempMasuX][self.tempMasuY] = self.sub_menu_item_dict[self.constriction_flag]["obj_flag"] - 1

    def RightClickEvent(self,event):
        if __debug__:
            pass
        pass

    def RightPressEvent(self,event):
        if __debug__:
            print("Pressed Pos:",event.x,event.y)

        self.RightPressPos_x = event.x
        self.RightPressPos_y = event.y

        pass

    def RightReleaseEvent(self,event):
        if __debug__:
            print("Release Pos:",event.x,event.y)
            print(self.map_chip)

        move_distance_x = self.RightPressPos_x - event.x
        move_distance_y = self.RightPressPos_y - event.y

        if __debug__:
            print("move distance:",move_distance_x,move_distance_y)
            print(self.masu_rect_list)

        self.startX += move_distance_x
        self.startY += move_distance_y

        self.DestroyAllMasu()
        self.SetMasu(self.masu_x,self.masu_y)

        if __debug__:
            print(self.map_chip)

        move_distance_x = 0
        move_distance_y = 0
        pass



    def SetButton(self,text,x,y,func):
        self.btn = tk.Button(self.canvas,text = text, width = 4, height = 2, bd=1, padx=0, pady=0, relief='ridge',command=func)
        self.btn.place(x=x, y=y)
        #self.btn.pack()
        self.sub_menu_button_list.append(self.btn)

    def SetButton(self,text,x,y,func,width = 4,height = 2):
        self.btn = tk.Button(self.canvas,text = text, width = width, height = height, bd=1, padx=0, pady=0, relief='ridge',command=func)
        self.btn.place(x=x, y=y)
        #self.btn.pack()
        self.sub_menu_button_list.append(self.btn)

    def SetButton(self, text, x, y, func, width=4, height=2, bg = "white"):
        self.btn = tk.Button(self.canvas, text=text, width=width, height=height, bd=1, padx=0, pady=0, relief='ridge',
                             command=func, bg = bg)
        self.btn.place(x=x, y=y)
        #self.btn.pack()
        self.sub_menu_button_list.append(self.btn)

    def ButtonEventTest(self):
        print("Button")


    def DrawSubMenuSingleItem(self,text,color = "red"):
        self.SetButton(text,x = self.sub_menu_button_start_position_x,y = self.sub_menu_button_start_position_y,
                       bg = color,
                       width = 8,
                       func = lambda f = text:self.CheckSubMenuItemNumber(f))

    def DrawSubMenuSingleItem(self,text):
        color = self.sub_menu_item_color[len(self.sub_menu_item_list)]
        self.SetButton(text,x = self.sub_menu_button_start_position_x,y = self.sub_menu_button_start_position_y,
                       bg = color,
                       width = 8,
                       func = lambda f = text:self.CheckSubMenuItemNumber(f))

    def OpenExcel(self,sheet_name):
        if(self.map_excel_file_full_name != ""):
            book = openpyxl.load_workbook(self.map_excel_file_full_name)
            sheet = book.get_sheet_by_name(sheet_name)
            return  {'sheet':sheet,'book':book}

    #Excelに登録したconfigを辞書リストに保存する
    def ReadSubMenuConfigFromExcel(self):
            sheet = self.OpenExcel("Submenu")['sheet']
            for i in range(2,30):
                obj_name = sheet.cell(row = i, column = 1).value
                obj_color = sheet.cell(row = i, column = 2).value
                obj_flag = sheet.cell(row = i, column = 3).value
                if(obj_name != None):
                    self.sub_menu_item_dict.append({'obj_name':obj_name,'obj_color':obj_color,'obj_flag':obj_flag})
                    self.sub_menu_item_color_can_be_choice.remove(obj_color)
                else:
                    break

            if __debug__:
                for i in range(len(self.sub_menu_item_dict)):
                    print(self.sub_menu_item_dict[i])


    def DrawSubMenuItem(self,x,y):
        if __debug__:
            print("length of sub_menu_item_dict",len(self.sub_menu_item_dict))
        for i in range(len(self.sub_menu_item_dict)):
            temp_obj_name = self.sub_menu_item_dict[i]['obj_name']
            temp_obj_color = self.sub_menu_item_dict[i]['obj_color']
            self.SetButton(temp_obj_name,x = x + (70 * int(i / 10)), y = y + self.sub_menu_button_padding_y * (i % 10),
                           bg = temp_obj_color,width = 8,
                           func = lambda f = temp_obj_name : self.CheckSubMenuItemNumber(f))




    def CheckSubMenuItemNumber(self, text):
        if __debug__:
            print("length of sub_men_button_list:",len(self.sub_menu_button_list))
        for i in range(len(self.sub_menu_button_list)):
            if(self.sub_menu_button_list[i]['text'] == text):
                self.constriction_flag = i
                if __debug__:
                    print("constriction_flag",i)
                    #print(self.sub_menu_item_color)
                break

    def SubMenuEvent(self):
        if __debug__:
            print("sub_menu_active",self.sub_menu_active)
            print("sub_button_list",self.sub_menu_button_list)
        if self.sub_menu_active == True:
            #self.canvas.delete(self.rect)
            #for i in (self.sub_menu_button_list):
                #print("type",type(i))
                #i.pack_forget()
            self.sub_menu_button_list = []
            self.sub_menu_active = False
        else:
            #self.DrawSubMenu()
            self.sub_menu_button_list = []
            self.DrawSubMenuItem(self.sub_menu_button_start_position_x,self.sub_menu_button_start_position_y)

            self.sub_menu_active = True

    def DestoryButton(self):
        for i in self.sub_menu_button_list:
            print(i)
            i.pack_forget()


    def DestroyAllMasu(self):
        for i in range(self.masu_x):
            for j in range(self.masu_y):
                self.canvas.delete(self.masu_rect_list[i][j])

    # 行追加
    def AddMasuColumn(self):
        self.DestroyAllMasu()
        self.masu_y += 1
        if __debug__:
            print(self.map_chip)
        for i in range(self.masu_x):
            self.map_chip[i].append(0)
        if __debug__:
            print(self.map_chip)
        self.SetMasu(self.masu_x,self.masu_y)
    # 行削除
    def RemoveMasuColumn(self):
        self.DestroyAllMasu()
        self.masu_y -= 1
        if __debug__:
            print(self.map_chip)
        for i in range(self.masu_x):
            self.map_chip[i].remove(self.map_chip[i][self.masu_y])
        if __debug__:
            print(self.map_chip)
        self.SetMasu(self.masu_x,self.masu_y)

    #
    def AddMasuRow(self):
        self.DestroyAllMasu()
        self.masu_x += 1
        if __debug__:
            print(self.map_chip)
        temp_list = []
        for i in range(self.masu_y):
            temp_list.append(0)
        self.map_chip.append(temp_list)
        if __debug__:
            print(self.map_chip)
        self.SetMasu(self.masu_x,self.masu_y)

    def RemoveMasuRow(self):
        self.DestroyAllMasu()
        self.masu_x -= 1
        if __debug__:
            print(self.map_chip)
        self.map_chip.remove(self.map_chip[self.masu_x])
        if __debug__:
            print(self.map_chip)
        self.SetMasu(self.masu_x,self.masu_y)

    def Set(self):
        if __debug__:
            self.test_btn = tk.Button(text = "行追加",command=self.AddMasuColumn)
            self.test_btn.place( x = self.sub_menu_button_start_position_x, y = 480,)

            self.test_btn = tk.Button(text = "列追加",command=self.AddMasuRow)
            self.test_btn.place( x = self.sub_menu_button_start_position_x + 50, y = 480,)

            self.test_btn = tk.Button(text = "行削除",command=self.RemoveMasuColumn)
            self.test_btn.place( x = self.sub_menu_button_start_position_x, y = 510,)

            self.test_btn = tk.Button(text = "列削除",command=self.RemoveMasuRow)
            self.test_btn.place( x = self.sub_menu_button_start_position_x + 50, y = 510,)

        self.canvas.bind('<ButtonPress-1>',self.LeftClickEvent)
        #self.canvas.bind('<ButtonPress-3>',self.RightClickEvent)
        self.canvas.bind('<ButtonPress-3>',self.RightPressEvent)
        self.canvas.bind('<ButtonPress-2>', self.ExcelForDirectX)
        self.canvas.bind('<ButtonRelease-3>',self.RightReleaseEvent)


    def ScaleUpMapChip(self,event):
        pass

    def ScaleDownMapChip(self,event):
        pass

    def DrawSubMenu(self):
        if self.sub_menu_active:
            self.rect = self.canvas.create_rectangle(800,20,900,400,fill = 'white')
            if __debug__:
                print("self.rect番号：",self.rect)
        else:
            pass

    def ReadMapInfoFromExcel(self):
        if(self.map_excel_file_full_name != ""):
            book = openpyxl.load_workbook(self.map_excel_file_full_name)
            active_sheet = book.active
            sheet = book.get_sheet_by_name("MapInfo")
            if __debug__:
                print("MapInfo Size X:",sheet.cell(row = 1, column = 2).value)
                print("MapInfo Size Y:",sheet.cell(row = 2, column = 2).value)

            self.masu_x = sheet.cell(row = 1, column= 2).value
            self.masu_y = sheet.cell(row = 2, column= 2).value
            #return (sheet.cell(row = 1,column=2).value,sheet.cell(row = 2,column=2).value)

    def AddNewSubMenuItem(self):
        pop_up_window = tk.Tk()
        pop_up_window.title("建物新規作成")
        pop_up_window.geometry("400x180")

        label_construction_name = tkinter.Label(pop_up_window,text = "建物の名前")
        label_construction_color = tkinter.Label(pop_up_window,text = "記号の色")
        label_construction_obj_adr = tkinter.Label(pop_up_window,text = "objファイルのアドレス")
        label_construction_name.place(x = 20, y = 20)
        label_construction_color.place(x = 20, y = 50)
        label_construction_obj_adr.place(x = 20, y = 80)


        #construction name
        construction_entry_box = tkinter.Entry(pop_up_window,width=40)
        construction_entry_box.insert(tkinter.END,"建物")
        construction_entry_box.place(x = 100, y = 20)

        #construction color
        construction_color_combobox = ttk.Combobox(pop_up_window,
                                                   height = len(self.sub_menu_item_color_can_be_choice),
                                                   values = self.sub_menu_item_color_can_be_choice)
        construction_color_combobox.place(x=100, y=50)
        construction_color_combobox.bind('<<ComboboxSelected>>',func = lambda f : print(construction_color_combobox.get()))

        #construction obg address
        def Choose_Obj_File():
            construction_obj_file_name = filedialog.askopenfilename(defaultextension=".obj")
            if __debug__:
                print(construction_obj_file_name)
            name_list = construction_obj_file_name.split('/')
            file_name = ""
            for i in range(-4,0):
                if i != -4:
                    file_name += ("/"+name_list[i])
                else:
                    file_name += name_list[i]
            if __debug__:
                print(file_name)
            construction_obj_file_adr_entry_box.insert(tkinter.END,file_name)

        construction_obj_file_adr_entry_box = tk.Entry(pop_up_window,width = 60)
        construction_obj_file_adr_entry_box.place(x = 20, y = 110)
        choice_btn = tk.Button(pop_up_window,text = "選択", command = Choose_Obj_File)
        choice_btn.pack()
        choice_btn.place(x = 150, y = 80)



        def GetPopUpWindowData():
            if __debug__:
                print("color:",construction_color_combobox.get())
                print("construction_name:",construction_entry_box.get())
            self.WriteSubMenuItemConfigToExcel(construction_entry_box.get(),
                                               construction_color_combobox.get(),
                                               construction_obj_file_adr_entry_box.get())
            pop_up_window.destroy()

            if __debug__:
                print("length of button list", len(self.sub_menu_button_list))

            if(self.sub_menu_active):
                #self.DrawSubMenu()
                self.sub_menu_button_list = []
                self.DrawSubMenuItem(self.sub_menu_button_start_position_x, self.sub_menu_button_start_position_y)


        pop_up_window_button = tkinter.Button(pop_up_window,text = "新規", command = GetPopUpWindowData)
        pop_up_window_button.pack()
        pop_up_window_button.place(x = 30, y = 140)

    def WriteSubMenuItemConfigToExcel(self,obj_name,obj_color,obj_adr):
        book = openpyxl.load_workbook(self.map_excel_file_full_name)
        sheet = book.get_sheet_by_name("Submenu")
        row = len(self.sub_menu_item_dict) + 2
        if __debug__:
            print("row:",row)
        sheet.cell(row = row, column = 1).value = obj_name
        sheet.cell(row = row, column = 2).value = obj_color
        sheet.cell(row = row, column = 3).value = self.SearchList(self.sub_menu_item_color,obj_color) + 1
        sheet.cell(row = row, column = 4).value = obj_adr
        self.sub_menu_item_dict.append({'obj_name': obj_name, 'obj_color': obj_color, 'obj_flag': self.SearchList(self.sub_menu_item_color,obj_color) + 1})
        self.sub_menu_item_color_can_be_choice.remove(obj_color)
        book.save(self.map_excel_file_full_name)


    def SearchList(self,list,obj):
        if __debug__:
            print("Search List")
            print(list)
            print(obj)
        for i in range(len(list)):
            if __debug__:
                print(list[i],obj)
            if list[i] == obj:
                if __debug__:
                    print("i:",i)
                return i

    def DrawMenu(self):
        self.menu_bar = tk.Menu(self.canvas)

        self.file_menu = tk.Menu(self.menu_bar, tearoff = False)
        self.menu_bar.add_cascade(label = "ファイル", menu = self.file_menu)

        self.file_menu.add_command(label = "新規",accelerator = "Ctrl+N")
        self.file_menu.add_command(label = "開く", command = self.OpenMapExcelFile, accelerator = "Ctrl+O")
        self.file_menu.add_command(label = "上書き保存", accelerator = "Ctrl+S",command = self.SaveMapChipToExcel)
        self.file_menu.add_command(label="名前を付けて保存", accelerator = "Ctrl+Alt+S")

        self.sub_menu = tk.Menu(self.menu_bar, tearoff = False)
        self.menu_bar.add_cascade(label = "サブメニュー", menu = self.sub_menu)

        self.sub_menu.add_command(label = "表示/隠す", command = self.SubMenuEvent)
        self.sub_menu.add_command(label = "新規", command = self.AddNewSubMenuItem)

        self.canvas.bind('<Control-Key-O>', self.OpenMapExcelFile)
        self.canvas.bind('<Control-Key-o>',self.OpenMapExcelFile)

    def OpenExplorer(self,event = 0):
        self.map_excel_file_full_name = filedialog.askopenfilename(defaultextension = ".xlsx")
        name_list = self.map_excel_file_full_name.split('/')
        self.map_excel_file_name = name_list[-1]
        #print(self.map_excel_file_name)

    def OpenMapExcelFile(self):
        self.OpenExplorer()
        self.ReadMapInfoFromExcel()
        self.ReadMapFromExcel()
        if(self.masu_x != -1 and self.masu_y != -1):
            self.SetMasu(self.masu_x, self.masu_y)
        self.ReadSubMenuConfigFromExcel()


    def ReadMapFromExcel(self):
        if (self.map_excel_file_full_name != ""):
            book = openpyxl.load_workbook(self.map_excel_file_full_name)
            active_sheet = book.active
            sheet = book.get_sheet_by_name("Map")
            for i in range(1,self.masu_x + 1):
                temp_list = []
                for j in range(1,self.masu_y + 1):
                    temp_value = sheet.cell(row = j, column= i).value
                    if __debug__:
                        print("cell value:",(j,i),":",temp_value)
                    if(temp_value == None):
                        temp_list.append(0)
                    else:
                        temp_list.append(temp_value)
                self.map_chip.append(temp_list)
            if __debug__:
                print(self.map_chip)
    def SetMinus1(self,x,y,width,height):
        for i in range(x,x + width):
            for j in range(y,y+height):
                self.map_chip[i][j] = -1



    def ExcelForDirectX(self,event):
        # ブロックInfo記録用辞書型リスト flag: width: height:
        temp_dict_list = []

        temp_construction = self.map_chip[0][0]
        if __debug__:
            print("temp_construction",temp_construction)

        # 建築ブロックの幅と深さ記録用
        temp_construction_width = 1
        temp_construction_height = 1

        # 探索位置の記録用
        temp_x = 0
        temp_y = 0
                        # 8
        for i in range(self.masu_y):
                            # 6
            for j in range(self.masu_x):
                if __debug__:
                    pass
                    #print(self.map_chip[i][j])
            #print("\n")
                # 横方向を軸して判定 (-1は探索済みのマスと表している) このマスの探索をskip
                if(self.map_chip[i][j] == -1):
                    pass

                # 未探索のマス
                else:
                    # 横方向の端っこにたどり着いた場合
                    if(j + 1 >= self.masu_x):
                        temp_x = j
                        temp_y = i
                        i += 1
                    # 次のコマも同じフラグなら
                    elif(self.map_chip[i][j + 1] == temp_construction):
                        # 建物ブロックの幅１伸ばせる
                        temp_construction_width += 1
                    # じゃない場合　今の位置を記録し　縦方向の探索開始
                    else:
                        temp_x = j
                        temp_y = i
                        i += 1

                        # 縦方向の判定 次のコマも同じフラグなら
                        if(i < self.masu_y):
                            if __debug__:
                                print("i:",i,"j:",j,"map_chip:",self.map_chip[i][j])
                            if(self.map_chip[i][j] == temp_construction):
                                # 建物ブロックの深さ１伸ばせる
                                temp_construction_height += 1
                            # このブロックの判定完了
                            else:
                                #辞書型のデータにまとめリストに保存
                                temp_dict_list.append({'flag':temp_construction,
                                                       'width':temp_construction_width,
                                                       'height':temp_construction_height})

                                # 探索済みのブロックを-1で埋る
                                self.SetMinus1()


                                if __debug__:
                                    print(temp_dict_list)

                                # 記録した位置まで戻す
                                i = temp_y
                                j = temp_x

                                # 　横方向の次のコマが探索対象となる
                                temp_construction = self.map_chip[i][j + 1]

                                # ブロック記録用の幅と深さのリセット
                                temp_construction_height = 1
                                temp_construction_width = 1
        if __debug__:
            print(temp_dict_list)










if __name__ == "__main__":
    root = tk.Tk()
    app = Mapedit(master = root)
    app.DrawMenu()
    #app.DrawSubMenu()
    app.Set()

    root.config(menu = app.menu_bar)
    app.mainloop()